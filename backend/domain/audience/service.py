import io
import json
import math
from typing import List
from sqlalchemy import func
from sqlalchemy.orm import Session
from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl

from domain.audience.models import ContactGroup, Contact
from domain.audience.schemas import GroupCreate, GroupUpdate, ContactCreate, GroupOut, PaginatedResponse


# ========== 客群操作 ==========
def list_groups(db: Session, user_id: int, search: str = "", page: int = 1, page_size: int = 20) -> PaginatedResponse:
    query = db.query(ContactGroup).filter(ContactGroup.user_id == user_id)
    if search:
        query = query.filter(ContactGroup.name.ilike(f"%{search}%"))

    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    groups = query.order_by(ContactGroup.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for g in groups:
        count = db.query(func.count(Contact.id)).filter(Contact.group_id == g.id).scalar()
        items.append(GroupOut(
            id=g.id, name=g.name, description=g.description,
            user_id=g.user_id, contact_count=count
        ))

    return PaginatedResponse(items=items, total=total, page=page, page_size=page_size, total_pages=total_pages)


def create_group(db: Session, data: GroupCreate, user_id: int) -> ContactGroup:
    group = ContactGroup(name=data.name, description=data.description, user_id=user_id)
    db.add(group)
    db.commit()
    db.refresh(group)
    return group


def update_group(db: Session, group_id: int, data: GroupUpdate, user_id: int) -> ContactGroup:
    group = _get_user_group(db, group_id, user_id)
    if data.name is not None:
        group.name = data.name
    if data.description is not None:
        group.description = data.description
    db.commit()
    db.refresh(group)
    return group


def delete_group(db: Session, group_id: int, user_id: int) -> dict:
    group = _get_user_group(db, group_id, user_id)
    db.delete(group)
    db.commit()
    return {"message": "客群已删除"}


# ========== 联系人操作 ==========
def list_contacts(db: Session, group_id: int, user_id: int, search: str = "", page: int = 1, page_size: int = 20) -> PaginatedResponse:
    _get_user_group(db, group_id, user_id)
    query = db.query(Contact).filter(Contact.group_id == group_id)
    if search:
        query = query.filter(
            (Contact.name.ilike(f"%{search}%")) | (Contact.email.ilike(f"%{search}%"))
        )

    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    rows = query.order_by(Contact.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    items = [{"id": c.id, "email": c.email, "name": c.name, "attributes": c.attributes, "group_id": c.group_id} for c in rows]

    return PaginatedResponse(items=items, total=total, page=page, page_size=page_size, total_pages=total_pages)


def create_contact(db: Session, data: ContactCreate, user_id: int) -> Contact:
    _get_user_group(db, data.group_id, user_id)
    contact = Contact(email=data.email, name=data.name, attributes=data.attributes, group_id=data.group_id)
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return contact


def delete_contact(db: Session, contact_id: int, user_id: int) -> dict:
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="联系人不存在")
    _get_user_group(db, contact.group_id, user_id)
    db.delete(contact)
    db.commit()
    return {"message": "联系人已删除"}


# ========== Excel 操作 ==========
def download_contacts_excel(db: Session, group_id: int, user_id: int) -> StreamingResponse:
    group = _get_user_group(db, group_id, user_id)
    contacts = db.query(Contact).filter(Contact.group_id == group_id).all()

    # Collect all attribute keys across contacts
    all_attr_keys = []
    for c in contacts:
        if c.attributes:
            try:
                attrs = json.loads(c.attributes)
                for k in attrs:
                    if k not in all_attr_keys:
                        all_attr_keys.append(k)
            except (json.JSONDecodeError, TypeError):
                pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "联系人"
    ws.append(["姓名", "邮箱"] + all_attr_keys)
    for c in contacts:
        attrs = {}
        if c.attributes:
            try:
                attrs = json.loads(c.attributes)
            except (json.JSONDecodeError, TypeError):
                pass
        row = [c.name or "", c.email] + [str(attrs.get(k, "")) for k in all_attr_keys]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={group.name}_contacts.xlsx"},
    )


def download_template_excel() -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "联系人模版"
    ws.append(["姓名", "邮箱", "company", "city"])
    ws.append(["张三", "zhangsan@example.com", "Acme Inc", "Shanghai"])
    ws.append(["李四", "lisi@example.com", "Tech Corp", "Beijing"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contact_template.xlsx"},
    )


def upload_contacts_excel(db: Session, group_id: int, user_id: int, file: UploadFile) -> dict:
    _get_user_group(db, group_id, user_id)

    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active

        # Read header row to get attribute column names
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        attr_keys = headers[2:]  # columns after 姓名, 邮箱

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0] or "").strip() if row[0] else ""
            email = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
            if not email:
                continue

            # Build attributes JSON from extra columns
            attrs = {}
            for i, key in enumerate(attr_keys):
                if key and len(row) > i + 2 and row[i + 2]:
                    attrs[key] = str(row[i + 2]).strip()

            contact = Contact(
                name=name,
                email=email,
                attributes=json.dumps(attrs, ensure_ascii=False) if attrs else None,
                group_id=group_id,
            )
            db.add(contact)
            count += 1
        db.commit()
        return {"message": f"成功导入 {count} 个联系人"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")


# ========== 内部工具 ==========
def _get_user_group(db: Session, group_id: int, user_id: int) -> ContactGroup:
    """获取属于指定用户的客群，不存在则抛异常"""
    group = db.query(ContactGroup).filter(
        ContactGroup.id == group_id, ContactGroup.user_id == user_id
    ).first()
    if not group:
        raise HTTPException(status_code=404, detail="客群不存在或无权操作")
    return group
