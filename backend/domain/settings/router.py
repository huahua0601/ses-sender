from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from core.database import get_db
from core.deps import require_admin, get_current_user
from domain.auth.models import User
from domain.settings import service

router = APIRouter(tags=["系统设置"])


@router.get("/admin/settings")
def get_settings(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    return service.get_all_settings(db)


@router.put("/admin/settings")
def save_settings(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    service.save_settings(db, data)
    return {"message": "配置已保存"}


@router.get("/admin/ai-models")
def list_ai_models(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    return service.get_ai_models(db)


@router.put("/admin/ai-models")
def save_ai_models(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    service.save_ai_models(db, data.get("models", []))
    return {"message": "模型列表已保存"}


@router.post("/admin/ai-models/test")
def test_ai_model(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    return service.test_ai_model(db, data)


@router.get("/ai-models/available")
def available_models(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """普通用户获取可用模型列表（按 Provider 分组）"""
    return service.get_available_models(db)


@router.post("/admin/settings/test-bedrock")
def test_ai_connection(data: dict = None, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    provider = (data or {}).get("ai_provider") or service.get_ai_provider(db)
    if provider == "openai_compatible":
        return service.test_openai_connection(db, data)
    return service.test_bedrock_connection(db, data)


@router.post("/admin/sql")
def execute_sql(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    """管理员执行 SQL 查询（仅支持 SELECT）"""
    from sqlalchemy import text
    sql = (data.get("sql") or "").strip()
    if not sql:
        raise HTTPException(status_code=400, detail="SQL 不能为空")

    sql_upper = sql.upper().lstrip()
    is_select = sql_upper.startswith("SELECT") or sql_upper.startswith("SHOW") or sql_upper.startswith("DESCRIBE") or sql_upper.startswith("EXPLAIN")

    allow_write = data.get("allow_write", False)
    if not is_select and not allow_write:
        raise HTTPException(status_code=400, detail="仅支持 SELECT/SHOW/DESCRIBE 查询。如需执行写操作请勾选「允许写操作」")

    try:
        result = db.execute(text(sql))
        if is_select or sql_upper.startswith("SHOW") or sql_upper.startswith("DESCRIBE") or sql_upper.startswith("EXPLAIN"):
            columns = list(result.keys()) if result.returns_rows else []
            rows = [dict(zip(columns, row)) for row in result.fetchall()] if columns else []
            return {"columns": columns, "rows": rows, "row_count": len(rows)}
        else:
            db.commit()
            return {"columns": [], "rows": [], "row_count": result.rowcount, "message": f"执行成功，影响 {result.rowcount} 行"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"SQL 执行错误: {str(e)}")


import os
from collections import deque

_LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "logs")


def _read_log_lines(lines: int = 200) -> list[str]:
    log_file = os.path.join(_LOG_DIR, "app.log")
    if not os.path.exists(log_file):
        return []
    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
        return list(deque(f, maxlen=lines))


@router.get("/admin/logs")
def get_logs(lines: int = Query(200, ge=10, le=10000), admin: User = Depends(require_admin)):
    content = _read_log_lines(lines)
    return {"lines": content, "total": len(content)}


@router.get("/admin/logs/download")
def download_logs(lines: int = Query(5000, ge=100, le=50000), admin: User = Depends(require_admin)):
    content = _read_log_lines(lines)

    def generate():
        for line in content:
            yield line

    from datetime import datetime
    filename = f"ses-sender-logs-{datetime.now().strftime('%Y%m%d-%H%M%S')}.log"
    return StreamingResponse(
        generate(),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========================
# 邮箱黑名单管理
# ========================

@router.get("/admin/blacklist")
def list_blacklist(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = Query(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from domain.sending.models import EmailBlacklist
    import math

    query = db.query(EmailBlacklist)
    if search:
        query = query.filter(EmailBlacklist.email.contains(search))
    total = query.count()
    rows = query.order_by(EmailBlacklist.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [{"id": r.id, "email": r.email, "reason": r.reason, "created_by": r.created_by, "created_at": r.created_at.isoformat() if r.created_at else None} for r in rows],
        "total": total,
        "page": page,
        "total_pages": math.ceil(total / page_size) if total else 1,
    }


@router.post("/admin/blacklist")
def add_blacklist(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    from domain.sending.models import EmailBlacklist
    from core import blacklist as cache

    email = (data.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="邮箱不能为空")

    existing = db.query(EmailBlacklist).filter(EmailBlacklist.email == email).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"{email} 已在黑名单中")

    record = EmailBlacklist(email=email, reason=data.get("reason", ""), created_by=admin.username)
    db.add(record)
    db.commit()
    cache.add(email)
    return {"message": f"已添加 {email} 到黑名单"}


@router.delete("/admin/blacklist/{item_id}")
def remove_blacklist(item_id: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    from domain.sending.models import EmailBlacklist
    from core import blacklist as cache

    record = db.query(EmailBlacklist).filter(EmailBlacklist.id == item_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    email = record.email
    db.delete(record)
    db.commit()
    cache.remove(email)
    return {"message": f"已从黑名单移除 {email}"}


@router.post("/admin/blacklist/batch-delete")
def batch_delete_blacklist(data: dict, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    from domain.sending.models import EmailBlacklist
    from core import blacklist as cache

    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="未选择记录")

    records = db.query(EmailBlacklist).filter(EmailBlacklist.id.in_(ids)).all()
    for r in records:
        cache.remove(r.email)
        db.delete(r)
    db.commit()
    return {"message": f"已删除 {len(records)} 条记录"}


@router.post("/admin/blacklist/upload")
async def upload_blacklist(file: UploadFile = File(...), admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    from domain.sending.models import EmailBlacklist
    from core import blacklist as cache

    contents = await file.read()
    filename = (file.filename or "").lower()
    emails = set()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    v = cell.strip().lower()
                    if "@" in v and "." in v and len(v) < 255:
                        emails.add(v)
    else:
        text = contents.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            line = line.strip().strip(",").strip(";").lower()
            if "@" in line and "." in line:
                parts = line.split(",")
                for p in parts:
                    p = p.strip()
                    if "@" in p and "." in p and len(p) < 255:
                        emails.add(p)

    if not emails:
        raise HTTPException(status_code=400, detail="未识别到有效邮箱地址")

    existing = set(r[0] for r in db.query(EmailBlacklist.email).filter(EmailBlacklist.email.in_(emails)).all())
    new_emails = emails - existing

    for email in new_emails:
        db.add(EmailBlacklist(email=email, reason="批量导入", created_by=admin.username))
    db.commit()
    cache.reload()

    return {"message": f"导入完成：{len(new_emails)} 个新增，{len(existing)} 个已存在", "added": len(new_emails), "skipped": len(existing)}


@router.get("/admin/blacklist/count")
def blacklist_count(admin: User = Depends(require_admin)):
    from core import blacklist as cache
    return {"count": cache.count()}


@router.get("/admin/blacklist/template")
def download_blacklist_template():
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "黑名单模板"
    ws.append(["email", "reason"])
    ws.append(["example@domain.com", "硬退信"])
    ws.append(["invalid@test.com", "无效邮箱"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="blacklist_template.xlsx"'},
    )
